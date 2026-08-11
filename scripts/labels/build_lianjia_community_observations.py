"""Rebuild purchased Lianjia transactions with correct price semantics.

The workbooks expose total transaction price in ten-thousand CNY and building
area in square metres.  This streaming builder computes CNY/m2, parses the deal
year, retains invalid rows with quality flags, and writes source-preserving
city fragments.  Raw workbooks and legacy labels are never modified.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import shutil
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import UTC, date, datetime
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "scripts"))
from urban_intervention.config.project import CITIES  # noqa: E402
from urban_intervention.data.paths import (  # noqa: E402
    OUTPUT_HOUSING_FUSION_DIR,
    RAW_LIANJIA_DIR,
    STAGING_LIANJIA_TRANSACTIONS_DIR,
)

RAW_DIR = RAW_LIANJIA_DIR
OUTPUT_DIR = STAGING_LIANJIA_TRANSACTIONS_DIR
REPORT_PATH = OUTPUT_HOUSING_FUSION_DIR / "lianjia_rebuild.json"
CHUNK_SIZE = 50_000


SCHEMA = pa.schema(
    [
        ("source_record_id", pa.string()),
        ("source_file", pa.string()),
        ("source_row", pa.int64()),
        ("city_key", pa.string()),
        ("district", pa.string()),
        ("community_name", pa.string()),
        ("community_name_normalized", pa.string()),
        ("community_key_exact", pa.string()),
        ("listing_date", pa.string()),
        ("deal_date", pa.string()),
        ("year", pa.int16()),
        ("total_price_10k_cny", pa.float64()),
        ("listing_price_10k_cny", pa.float64()),
        ("deal_cycle_days", pa.float64()),
        ("price_adjustment_count", pa.float64()),
        ("showing_count", pa.float64()),
        ("follower_count", pa.float64()),
        ("view_count", pa.float64()),
        ("building_area_m2", pa.float64()),
        ("usable_area_m2", pa.float64()),
        ("unit_price_cny_m2", pa.float64()),
        ("layout", pa.string()),
        ("bedroom_count", pa.int16()),
        ("living_room_count", pa.int16()),
        ("floor_raw", pa.string()),
        ("layout_structure", pa.string()),
        ("building_type", pa.string()),
        ("orientation", pa.string()),
        ("built_year_raw", pa.string()),
        ("built_year_mid", pa.float64()),
        ("decoration", pa.string()),
        ("building_structure", pa.string()),
        ("heating", pa.string()),
        ("elevator_ratio", pa.string()),
        ("tenure_years", pa.string()),
        ("elevator", pa.string()),
        ("transaction_ownership", pa.string()),
        ("property_use", pa.string()),
        ("property_age", pa.string()),
        ("property_owner", pa.string()),
        ("hedonic_feature_count", pa.int16()),
        ("lon", pa.float64()),
        ("lat", pa.float64()),
        ("year_valid", pa.bool_()),
        ("price_valid", pa.bool_()),
        ("coordinate_valid", pa.bool_()),
        ("community_valid", pa.bool_()),
        ("is_valid", pa.bool_()),
        ("quality_flags", pa.string()),
    ]
)


def normalize_name(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    return re.sub(r"[\s\-—_·•,，、。.;；:：()（）\[\]【】]+", "", text)


def _to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        result = float(str(value).replace(",", "").strip())
        return result if math.isfinite(result) else None
    except (TypeError, ValueError):
        return None


def _date_text(value: object) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return str(value or "").strip()


def _text(value: object) -> str:
    return str(value or "").strip()


def _parse_layout(value: object) -> tuple[int | None, int | None]:
    text = _text(value)
    bedroom = re.search(r"(\d+)\s*室", text)
    living_room = re.search(r"(\d+)\s*厅", text)
    return (
        int(bedroom.group(1)) if bedroom else None,
        int(living_room.group(1)) if living_room else None,
    )


def _parse_built_year_mid(value: object) -> float | None:
    years = [int(year) for year in re.findall(r"(?:19|20)\d{2}", _text(value))]
    years = [year for year in years if 1900 <= year <= 2030]
    return float(sum(years) / len(years)) if years else None


def _parse_year(deal_date: object, fallback: object) -> int | None:
    if isinstance(deal_date, (datetime, date)):
        return int(deal_date.year)
    match = re.search(r"(19|20)\d{2}", str(deal_date or ""))
    if match:
        return int(match.group(0))
    try:
        year = int(float(fallback))
        return year
    except (TypeError, ValueError):
        return None


def _city_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for key, cfg in CITIES.items():
        name = str(cfg["name"]).replace("市", "").strip()
        lookup[name] = key
        lookup[str(cfg["name"]).strip()] = key
    return lookup


CITY_LOOKUP = _city_lookup()


def _city_key(value: object) -> str | None:
    text = str(value or "").strip()
    if text in CITY_LOOKUP:
        return CITY_LOOKUP[text]
    short = text.replace("市", "").strip()
    if short in CITY_LOOKUP:
        return CITY_LOOKUP[short]
    for name, key in CITY_LOOKUP.items():
        if len(name) >= 2 and (name in text or text in name):
            return key
    return None


def _year_range(city_key: str) -> tuple[int, int]:
    if city_key == "beijing":
        return 2010, 2023
    if city_key == "chongqing":
        return 2015, 2023
    return 2008, 2023


def _distinct_workbooks() -> list[Path]:
    distinct: dict[tuple[str, int], Path] = {}
    for path in sorted(RAW_DIR.rglob("*.xlsx")):
        distinct.setdefault((path.name, path.stat().st_size), path)
    return sorted(distinct.values())


def _column_indices(headers: list[object]) -> dict[str, int]:
    names = {str(value).strip(): index for index, value in enumerate(headers)}

    def first(*candidates: str) -> int | None:
        for candidate in candidates:
            if candidate in names:
                return names[candidate]
        return None

    result = {
        "city": first("城市"),
        "district": first("区域"),
        "community": first("小区"),
        "listing_date": first("挂牌时间", "挂牌日期"),
        "deal_date": first("成交日期", "成交时间"),
        "year": first("成交年份", "年份"),
        "total_price": first("成交价（万）", "成交价(万)", "成交价"),
        "listing_price": first("挂牌价格（万）", "挂牌价格(万)", "挂牌价格"),
        "deal_cycle": first("成交周期（天）", "成交周期(天)", "成交周期"),
        "price_adjustments": first("调价（次）", "调价(次)", "调价"),
        "showings": first("带看（次）", "带看(次)", "带看"),
        "followers": first("关注（人）", "关注(人)", "关注"),
        "views": first("浏览（次）", "浏览(次)", "浏览"),
        "area": first("建筑面积（㎡）", "建筑面积(㎡)", "建筑面积"),
        "usable_area": first("套内面积（㎡）", "套内面积(㎡)", "套内面积"),
        "layout": first("房屋户型"),
        "floor": first("所在楼层"),
        "layout_structure": first("户型结构"),
        "building_type": first("建筑类型"),
        "orientation": first("房屋朝向"),
        "built_year": first("建成年代"),
        "decoration": first("装修情况"),
        "building_structure": first("建筑结构"),
        "heating": first("供暖方式"),
        "elevator_ratio": first("梯户比例"),
        "tenure": first("产权年限"),
        "elevator": first("配备电梯"),
        "transaction_ownership": first("交易权属"),
        "property_use": first("房屋用途"),
        "property_age": first("房屋年限"),
        "property_owner": first("房权所属"),
        "lon": first("经度"),
        "lat": first("纬度"),
    }
    missing = [
        key
        for key in ("city", "community", "total_price", "area", "lon", "lat")
        if result[key] is None
    ]
    if missing:
        raise ValueError(f"missing required columns {missing}; headers={headers}")
    return {key: int(value) for key, value in result.items() if value is not None}


class FragmentWriters:
    def __init__(self, workbook_key: str):
        self.final_root = OUTPUT_DIR / workbook_key
        self.root = OUTPUT_DIR / f".{workbook_key}.tmp"
        if self.root.exists():
            shutil.rmtree(self.root)
        self.root.mkdir(parents=True, exist_ok=True)
        self.writers: dict[str, pq.ParquetWriter] = {}

    def write(self, city_key: str, records: list[dict]) -> None:
        if not records:
            return
        table = pa.Table.from_pylist(records, schema=SCHEMA)
        if city_key not in self.writers:
            self.writers[city_key] = pq.ParquetWriter(
                self.root / f"{city_key}.parquet", SCHEMA, compression="zstd"
            )
        self.writers[city_key].write_table(table)

    def close(self) -> None:
        for writer in self.writers.values():
            writer.close()

    def commit(self) -> None:
        backup = OUTPUT_DIR / f".{self.final_root.name}.backup"
        if backup.exists():
            shutil.rmtree(backup)
        if self.final_root.exists():
            self.final_root.replace(backup)
        self.root.replace(self.final_root)
        if backup.exists():
            shutil.rmtree(backup)


def process_workbook(path: Path) -> dict:
    workbook_key = hashlib.sha1(str(path.relative_to(RAW_DIR)).encode("utf-8")).hexdigest()[:12]
    writers = FragmentWriters(workbook_key)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = list(next(iterator))
    indices = _column_indices(headers)

    buffers: dict[str, list[dict]] = defaultdict(list)
    counts: Counter = Counter()
    valid_by_city: Counter = Counter()
    rows_by_city: Counter = Counter()

    def flush(city_key: str) -> None:
        writers.write(city_key, buffers[city_key])
        buffers[city_key].clear()

    for source_row, values in enumerate(iterator, start=2):
        counts["source_rows"] += 1

        def get(key, values=values):
            return values[indices[key]] if key in indices and indices[key] < len(values) else None

        city_key = _city_key(get("city"))
        if city_key is None:
            counts["non_research_city"] += 1
            continue
        rows_by_city[city_key] += 1
        community = str(get("community") or "").strip()
        normalized = normalize_name(community)
        deal_date = get("deal_date")
        year = _parse_year(deal_date, get("year"))
        total_price = _to_float(get("total_price"))
        listing_price = _to_float(get("listing_price"))
        area = _to_float(get("area"))
        usable_area = _to_float(get("usable_area"))
        lon = _to_float(get("lon"))
        lat = _to_float(get("lat"))
        unit_price = total_price * 10000.0 / area if total_price and area and area > 0 else None
        layout = _text(get("layout"))
        bedroom_count, living_room_count = _parse_layout(layout)
        built_year_raw = _text(get("built_year"))
        built_year_mid = _parse_built_year_mid(built_year_raw)
        feature_values = [
            bedroom_count,
            living_room_count,
            _text(get("floor")),
            built_year_mid,
            _text(get("decoration")),
            _text(get("orientation")),
            _text(get("building_type")),
            _text(get("elevator")),
            usable_area,
        ]
        hedonic_feature_count = sum(
            value is not None and (not isinstance(value, str) or bool(value))
            for value in feature_values
        )

        min_year, max_year = _year_range(city_key)
        year_valid = year is not None and min_year <= year <= max_year
        price_valid = unit_price is not None and 500 <= unit_price <= 500_000
        coordinate_valid = (
            lon is not None and lat is not None and 70 <= lon <= 140 and 15 <= lat <= 55
        )
        community_valid = len(normalized) >= 2
        flags = []
        if not year_valid:
            flags.append("invalid_year")
        if not price_valid:
            flags.append("invalid_price")
        if not coordinate_valid:
            flags.append("invalid_coordinate")
        if not community_valid:
            flags.append("invalid_community")
        is_valid = not flags
        if is_valid:
            valid_by_city[city_key] += 1
            counts["valid_rows"] += 1
        else:
            for flag in flags:
                counts[flag] += 1

        record_id = hashlib.sha1(f"{path.relative_to(RAW_DIR)}|{source_row}".encode()).hexdigest()
        buffers[city_key].append(
            {
                "source_record_id": record_id,
                "source_file": str(path.relative_to(ROOT)),
                "source_row": source_row,
                "city_key": city_key,
                "district": str(get("district") or "").strip(),
                "community_name": community,
                "community_name_normalized": normalized,
                "community_key_exact": f"{city_key}|{normalized}",
                "listing_date": _date_text(get("listing_date")),
                "deal_date": _date_text(deal_date),
                "year": int(year) if year is not None and -32768 <= year <= 32767 else -1,
                "total_price_10k_cny": total_price,
                "listing_price_10k_cny": listing_price,
                "deal_cycle_days": _to_float(get("deal_cycle")),
                "price_adjustment_count": _to_float(get("price_adjustments")),
                "showing_count": _to_float(get("showings")),
                "follower_count": _to_float(get("followers")),
                "view_count": _to_float(get("views")),
                "building_area_m2": area,
                "usable_area_m2": usable_area,
                "unit_price_cny_m2": unit_price,
                "layout": layout,
                "bedroom_count": bedroom_count,
                "living_room_count": living_room_count,
                "floor_raw": _text(get("floor")),
                "layout_structure": _text(get("layout_structure")),
                "building_type": _text(get("building_type")),
                "orientation": _text(get("orientation")),
                "built_year_raw": built_year_raw,
                "built_year_mid": built_year_mid,
                "decoration": _text(get("decoration")),
                "building_structure": _text(get("building_structure")),
                "heating": _text(get("heating")),
                "elevator_ratio": _text(get("elevator_ratio")),
                "tenure_years": _text(get("tenure")),
                "elevator": _text(get("elevator")),
                "transaction_ownership": _text(get("transaction_ownership")),
                "property_use": _text(get("property_use")),
                "property_age": _text(get("property_age")),
                "property_owner": _text(get("property_owner")),
                "hedonic_feature_count": hedonic_feature_count,
                "lon": lon,
                "lat": lat,
                "year_valid": year_valid,
                "price_valid": price_valid,
                "coordinate_valid": coordinate_valid,
                "community_valid": community_valid,
                "is_valid": is_valid,
                "quality_flags": ";".join(flags),
            }
        )
        if len(buffers[city_key]) >= CHUNK_SIZE:
            flush(city_key)

    for city_key in list(buffers):
        flush(city_key)
    writers.close()
    writers.commit()
    workbook.close()
    return {
        "source_file": str(path.relative_to(ROOT)),
        "workbook_key": workbook_key,
        "counts": dict(counts),
        "rows_by_city": dict(rows_by_city),
        "valid_by_city": dict(valid_by_city),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", help="Optional source workbook file name substring")
    args = parser.parse_args()
    workbooks = _distinct_workbooks()
    if args.workbook:
        workbooks = [path for path in workbooks if args.workbook in path.name]
    if not workbooks:
        raise FileNotFoundError("No matching Lianjia workbook")

    results = []
    for index, path in enumerate(workbooks, start=1):
        print(f"[{index}/{len(workbooks)}] {path.relative_to(ROOT)}", flush=True)
        result = process_workbook(path)
        results.append(result)
        print(json.dumps(result["counts"], ensure_ascii=False), flush=True)

    report = {
        "created_at": datetime.now(UTC).isoformat(),
        "source_files_modified": False,
        "legacy_labels_modified": False,
        "formula": "unit_price_cny_m2 = total_price_10k_cny * 10000 / building_area_m2",
        "workbooks": results,
    }
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"saved {REPORT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
