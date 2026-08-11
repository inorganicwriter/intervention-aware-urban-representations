"""Import authorized housing exports into the canonical observation contract."""

from __future__ import annotations

import hashlib
import json
import re
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import yaml
from openpyxl import load_workbook

from urban_intervention.config.project import CITIES

SCHEMA_VERSION = "housing_observation_v2"
PIPELINE_VERSION = "housing_authorized_import_v1"
TRANSACTION_DEDUP_COLUMNS = (
    "city_key",
    "community_name",
    "deal_date",
    "unit_price_cny_m2",
    "total_price_10k_cny",
    "building_area_m2",
    "lon",
    "lat",
)
CANONICAL_COLUMNS = [
    "source_record_id",
    "source_platform",
    "acquisition_method",
    "batch_id",
    "city_key",
    "district",
    "community_name",
    "community_id",
    "observation_type",
    "price_type",
    "spatial_unit",
    "temporal_unit",
    "unit",
    "source_snapshot_date",
    "deal_date",
    "unit_price_cny_m2",
    "total_price_10k_cny",
    "building_area_m2",
    "lon",
    "lat",
    "layout",
    "bedroom_count",
    "floor_raw",
    "built_year",
    "decoration",
    "property_type",
    "source_url",
    "source_page_id",
    "source_file_sha256",
    "raw_row_number",
    "quality_flags",
    "pipeline_version",
]
NUMERIC_COLUMNS = [
    "unit_price_cny_m2",
    "total_price_10k_cny",
    "building_area_m2",
    "lon",
    "lat",
    "bedroom_count",
    "built_year",
]
REQUIRED_METADATA = {
    "batch_id",
    "source_platform",
    "acquisition_method",
    "price_type",
    "observation_type",
    "temporal_unit",
    "spatial_unit",
    "unit",
    "source_snapshot_date",
}


from urban_intervention.utils import sha256_file  # noqa: E402


def _load_table(
    path: Path,
    columns: list[str] | None = None,
    sheet_name: str | int | None = None,
) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path, encoding="utf-8-sig", low_memory=False, usecols=columns)
    if suffix == ".parquet":
        return pd.read_parquet(path, columns=columns)
    if suffix == ".dta":
        frame = pd.read_stata(path, convert_categoricals=False)
        return frame if columns is None else frame[columns]
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path, sheet_name=sheet_name or 0, usecols=columns)
    raise ValueError(f"Unsupported housing import format: {path.suffix}")


def load_mapping(path: Path) -> dict[str, Any]:
    payload = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("Housing import mapping must be a YAML object")
    metadata = payload.get("metadata", {})
    missing = REQUIRED_METADATA - set(metadata)
    if missing:
        raise ValueError(f"Mapping metadata missing {sorted(missing)}")
    batch_id = str(metadata["batch_id"])
    if not re.fullmatch(r"[a-zA-Z0-9][a-zA-Z0-9_.-]{0,99}", batch_id):
        raise ValueError("batch_id must be a safe 1-100 character identifier")
    columns = payload.get("columns", {})
    unknown = set(columns) - set(CANONICAL_COLUMNS)
    if unknown:
        raise ValueError(f"Unknown canonical mapping fields: {sorted(unknown)}")
    date_parts = payload.get("date_parts", {})
    unknown_date_targets = set(date_parts) - {"deal_date", "source_snapshot_date"}
    if unknown_date_targets:
        raise ValueError(f"Unsupported date_parts targets: {sorted(unknown_date_targets)}")
    for target, spec in date_parts.items():
        if not isinstance(spec, dict) or "year_column" not in spec:
            raise ValueError(f"date_parts.{target} requires year_column")
    date_formats = payload.get("date_formats", {})
    unknown_date_format_targets = set(date_formats) - {
        "deal_date",
        "source_snapshot_date",
    }
    if unknown_date_format_targets:
        raise ValueError(f"Unsupported date_formats targets: {sorted(unknown_date_format_targets)}")
    for target, spec in date_formats.items():
        if not isinstance(spec, dict) or not {"source_column", "format"} <= set(spec):
            raise ValueError(f"date_formats.{target} requires source_column and format")
    return payload


def _date_part_source_columns(mapping: dict[str, Any]) -> set[str]:
    columns: set[str] = set()
    for spec in mapping.get("date_parts", {}).values():
        columns.add(str(spec["year_column"]))
        if spec.get("month_column"):
            columns.add(str(spec["month_column"]))
        if spec.get("day_column"):
            columns.add(str(spec["day_column"]))
    for spec in mapping.get("date_formats", {}).values():
        columns.add(str(spec["source_column"]))
    return columns


def _quality_flags(frame: pd.DataFrame) -> pd.Series:
    flags: list[list[str]] = [[] for _ in range(len(frame))]

    def add(mask: pd.Series, label: str) -> None:
        for position in np.flatnonzero(mask.fillna(False).to_numpy()):
            flags[int(position)].append(label)

    add(frame["city_key"].isna() | ~frame["city_key"].isin(CITIES), "invalid_city")
    add(
        frame["unit_price_cny_m2"].notna() & (frame["unit_price_cny_m2"] <= 0),
        "nonpositive_unit_price",
    )
    add(
        frame["building_area_m2"].notna() & (frame["building_area_m2"] <= 0),
        "nonpositive_area",
    )
    add(frame["lon"].notna() & ~frame["lon"].between(-180, 180), "invalid_lon")
    add(frame["lat"].notna() & ~frame["lat"].between(-90, 90), "invalid_lat")
    transaction = frame["price_type"].eq("transaction")
    add(transaction & frame["deal_date"].isna(), "missing_deal_date")
    missing_transaction_price = transaction & frame["unit_price_cny_m2"].isna()
    add(missing_transaction_price, "missing_transaction_unit_price")
    listing = frame["price_type"].isin(["listing", "platform_estimate"])
    add(listing & frame["source_snapshot_date"].isna(), "missing_snapshot_date")

    formula_ready = (
        transaction
        & frame["unit_price_cny_m2"].gt(0)
        & frame["total_price_10k_cny"].gt(0)
        & frame["building_area_m2"].gt(0)
    )
    derived_price = frame["total_price_10k_cny"] * 10_000 / frame["building_area_m2"]
    relative_error = (frame["unit_price_cny_m2"] - derived_price).abs() / derived_price
    add(formula_ready & relative_error.gt(0.01), "price_formula_mismatch_gt_1pct")

    key_ready = (
        transaction
        & frame["city_key"].notna()
        & frame["community_name"].notna()
        & frame["deal_date"].notna()
        & frame["unit_price_cny_m2"].notna()
        & frame["building_area_m2"].notna()
        & frame["lon"].notna()
        & frame["lat"].notna()
    )
    duplicate_transaction = (
        frame.duplicated(list(TRANSACTION_DEDUP_COLUMNS), keep=False) & key_ready
    )
    add(duplicate_transaction, "duplicate_transaction_key")
    return pd.Series([";".join(values) for values in flags], index=frame.index)


def _append_source_quality_rules(
    quality_flags: pd.Series,
    source: pd.DataFrame,
    rules: list[dict[str, Any]],
) -> pd.Series:
    result = quality_flags.copy()
    for rule in rules:
        source_column = str(rule["source_column"])
        if source_column not in source:
            raise ValueError(f"Quality rule source column is missing: {source_column}")
        operator = str(rule.get("operator", "equals"))
        if operator != "equals":
            raise ValueError(f"Unsupported quality rule operator: {operator}")
        mask = source[source_column].astype("string").eq(str(rule["value"]))
        flag = str(rule["flag"])
        result.loc[mask] = result.loc[mask].map(
            lambda value, flag=flag: ";".join(filter(None, [value, flag]))
        )
    return result


def normalize_authorized_export(
    source: pd.DataFrame,
    mapping: dict[str, Any],
    source_sha256: str,
    raw_row_start: int = 2,
) -> pd.DataFrame:
    metadata = mapping["metadata"]
    column_map = mapping.get("columns", {})
    missing_source_columns = sorted(set(column_map.values()) - set(source.columns))
    if missing_source_columns:
        raise ValueError(f"Input is missing mapped columns: {missing_source_columns}")

    result = pd.DataFrame(index=source.index)
    for canonical, source_column in column_map.items():
        result[canonical] = source[source_column]
    for target, spec in mapping.get("date_parts", {}).items():
        year = pd.to_numeric(source[str(spec["year_column"])], errors="coerce")
        if spec.get("month_column"):
            month = pd.to_numeric(source[str(spec["month_column"])], errors="coerce")
        else:
            month = pd.Series(spec.get("month", 1), index=source.index)
        if spec.get("day_column"):
            day = pd.to_numeric(source[str(spec["day_column"])], errors="coerce")
        else:
            day = pd.Series(spec.get("day", 1), index=source.index)
        result[target] = pd.to_datetime({"year": year, "month": month, "day": day}, errors="coerce")
    for target, spec in mapping.get("date_formats", {}).items():
        source_values = source[str(spec["source_column"])].astype("string").str.strip()
        # Stata commonly stores YYYYMM codes as floats, which pandas exposes as
        # strings such as "202001.0".  Removing only that terminal decimal keeps
        # the formatter explicit without coercing arbitrary dates.
        source_values = source_values.str.replace(r"\.0$", "", regex=True)
        result[target] = pd.to_datetime(
            source_values,
            format=str(spec["format"]),
            errors="coerce",
        )
    for column in CANONICAL_COLUMNS:
        if column not in result:
            result[column] = pd.NA
    for key, value in metadata.items():
        if key in CANONICAL_COLUMNS:
            result[key] = value

    city_map = {str(key): str(value) for key, value in mapping.get("city_key_map", {}).items()}
    result["city_key"] = result["city_key"].astype("string").str.strip().replace(city_map)
    for column in NUMERIC_COLUMNS:
        result[column] = pd.to_numeric(result[column], errors="coerce")
    result["deal_date"] = pd.to_datetime(result["deal_date"], errors="coerce")
    result["source_snapshot_date"] = pd.to_datetime(result["source_snapshot_date"], errors="coerce")

    derived = (
        result["unit_price_cny_m2"].isna()
        & result["total_price_10k_cny"].notna()
        & result["building_area_m2"].gt(0)
    )
    result.loc[derived, "unit_price_cny_m2"] = (
        result.loc[derived, "total_price_10k_cny"]
        * 10_000
        / result.loc[derived, "building_area_m2"]
    )

    result["raw_row_number"] = np.arange(raw_row_start, raw_row_start + len(result), dtype=np.int64)
    result["source_file_sha256"] = source_sha256
    result["pipeline_version"] = PIPELINE_VERSION
    result["source_record_id"] = result["source_record_id"].astype("string")
    missing_ids = result["source_record_id"].isna() | result["source_record_id"].str.strip().eq("")
    if missing_ids.any():
        generated = [
            hashlib.sha256(
                f"{metadata['source_platform']}|{metadata['batch_id']}|{source_sha256}|{row}".encode()
            ).hexdigest()[:24]
            for row in result.loc[missing_ids, "raw_row_number"]
        ]
        result.loc[missing_ids, "source_record_id"] = generated
    result["quality_flags"] = _quality_flags(result)
    result["quality_flags"] = _append_source_quality_rules(
        result["quality_flags"], source, mapping.get("quality_flag_rules", [])
    )
    duplicate_ids = result.duplicated(["source_platform", "source_record_id"], keep=False)
    result.loc[duplicate_ids, "quality_flags"] = result.loc[duplicate_ids, "quality_flags"].map(
        lambda value: ";".join(filter(None, [value, "duplicate_source_record_id"]))
    )
    return result[CANONICAL_COLUMNS].reset_index(drop=True)


def _append_flag_once(value: Any, flag: str) -> str:
    existing = [part for part in str(value or "").split(";") if part]
    if flag not in existing:
        existing.append(flag)
    return ";".join(existing)


def import_large_xlsx_authorized_export(
    input_path: Path,
    mapping_path: Path,
    raw_root: Path,
    staging_root: Path,
    chunk_rows: int = 25_000,
) -> tuple[Path, Path]:
    """Stream a large workbook in bounded-memory chunks before finalizing Parquet."""
    if chunk_rows <= 0:
        raise ValueError("chunk_rows must be positive")
    input_path = input_path.resolve()
    if input_path.suffix.lower() != ".xlsx":
        raise ValueError("Large streaming import currently supports .xlsx only")
    mapping = load_mapping(mapping_path)
    batch_id = str(mapping["metadata"]["batch_id"])
    raw_batch = raw_root / batch_id
    staging_batch = staging_root / batch_id
    parts_dir = staging_root / f".{batch_id}.parts"
    if raw_batch.exists() or staging_batch.exists() or parts_dir.exists():
        raise FileExistsError(f"Housing import batch already exists: {batch_id}")

    source_hash = sha256_file(input_path)
    selected_columns = set(mapping.get("columns", {}).values())
    selected_columns.update(_date_part_source_columns(mapping))
    selected_columns.update(
        str(rule["source_column"]) for rule in mapping.get("quality_flag_rules", [])
    )
    parts_dir.mkdir(parents=True)
    workbook = load_workbook(input_path, read_only=True, data_only=True, keep_links=False)
    sheet_name = mapping.get("sheet_name") or workbook.sheetnames[0]
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    missing = sorted(selected_columns - set(headers))
    if missing:
        workbook.close()
        raise ValueError(f"Input is missing mapped columns: {missing}")
    positions = {name: headers.index(name) for name in selected_columns}

    part_paths: list[Path] = []
    key_hashes: list[np.ndarray] = []
    key_ready_parts: list[np.ndarray] = []
    id_hashes: list[np.ndarray] = []
    buffer: dict[str, list[object]] = {name: [] for name in selected_columns}
    processed = 0

    def flush() -> None:
        nonlocal processed, buffer
        if not next(iter(buffer.values()), []):
            return
        source = pd.DataFrame(buffer)
        normalized = normalize_authorized_export(
            source, mapping, source_hash, raw_row_start=processed + 2
        )
        ready = (
            normalized["price_type"].eq("transaction")
            & normalized["city_key"].notna()
            & normalized["community_name"].notna()
            & normalized["deal_date"].notna()
            & normalized["unit_price_cny_m2"].notna()
            & normalized["building_area_m2"].notna()
            & normalized["lon"].notna()
            & normalized["lat"].notna()
        )
        key_hashes.append(
            pd.util.hash_pandas_object(
                normalized[list(TRANSACTION_DEDUP_COLUMNS)], index=False
            ).to_numpy(dtype=np.uint64)
        )
        key_ready_parts.append(ready.to_numpy(dtype=bool))
        id_hashes.append(
            pd.util.hash_pandas_object(normalized["source_record_id"], index=False).to_numpy(
                dtype=np.uint64
            )
        )
        part_path = parts_dir / f"part-{len(part_paths):05d}.parquet"
        normalized.to_parquet(part_path, index=False)
        part_paths.append(part_path)
        processed += len(normalized)
        buffer.clear()
        buffer.update({name: [] for name in selected_columns})

    for values in rows:
        for name, position in positions.items():
            buffer[name].append(values[position] if position < len(values) else None)
        if len(next(iter(buffer.values()))) >= chunk_rows:
            flush()
    flush()
    workbook.close()

    all_hashes = np.concatenate(key_hashes) if key_hashes else np.array([], dtype=np.uint64)
    all_ready = np.concatenate(key_ready_parts) if key_ready_parts else np.array([], dtype=bool)
    duplicate_key = pd.Series(all_hashes).duplicated(keep=False).to_numpy() & all_ready
    all_id_hashes = np.concatenate(id_hashes) if id_hashes else np.array([], dtype=np.uint64)
    duplicate_source_id = pd.Series(all_id_hashes).duplicated(keep=False).to_numpy()

    raw_batch.mkdir(parents=True)
    staging_batch.mkdir(parents=True)
    output_path = staging_batch / "housing_observations.parquet"
    writer: pq.ParquetWriter | None = None
    quality_counts: Counter[str] = Counter()
    offset = 0
    try:
        for part_path in part_paths:
            frame = pd.read_parquet(part_path)
            part_duplicate = duplicate_key[offset : offset + len(frame)]
            if part_duplicate.any():
                frame.loc[part_duplicate, "quality_flags"] = frame.loc[
                    part_duplicate, "quality_flags"
                ].map(lambda value: _append_flag_once(value, "duplicate_transaction_key"))
            part_dup_id = duplicate_source_id[offset : offset + len(frame)]
            if part_dup_id.any():
                frame.loc[part_dup_id, "quality_flags"] = frame.loc[
                    part_dup_id, "quality_flags"
                ].map(lambda value: _append_flag_once(value, "duplicate_source_record_id"))
            quality_counts.update(frame["quality_flags"].value_counts().to_dict())
            table = pa.Table.from_pandas(frame, preserve_index=False)
            if writer is None:
                writer = pq.ParquetWriter(output_path, table.schema)
            else:
                table = table.cast(writer.schema)
            writer.write_table(table)
            offset += len(frame)
    finally:
        if writer is not None:
            writer.close()

    raw_copy = raw_batch / input_path.name
    shutil.copy2(input_path, raw_copy)
    mapping_copy = raw_batch / "import_mapping.yaml"
    shutil.copy2(mapping_path, mapping_copy)
    manifest = {
        "schema": "housing_authorized_import_manifest_v1",
        "observation_schema": SCHEMA_VERSION,
        "pipeline_version": PIPELINE_VERSION,
        "import_mode": "streaming_xlsx",
        "chunk_rows": chunk_rows,
        "batch_id": batch_id,
        "metadata": mapping["metadata"],
        "input_file": input_path.name,
        "input_sha256": source_hash,
        "input_rows": processed,
        "output_rows": processed,
        "quality_flag_counts": dict(quality_counts),
        "output_file": str(output_path),
    }
    manifest_path = staging_batch / "import_manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    shutil.rmtree(parts_dir)
    return output_path, manifest_path


def resume_large_xlsx_authorized_export(
    input_path: Path,
    mapping_path: Path,
    raw_root: Path,
    staging_root: Path,
) -> tuple[Path, Path]:
    """Finalize complete streaming parts left by an interrupted import."""
    input_path = input_path.resolve()
    mapping = load_mapping(mapping_path)
    batch_id = str(mapping["metadata"]["batch_id"])
    raw_batch = raw_root / batch_id
    staging_batch = staging_root / batch_id
    parts_dir = staging_root / f".{batch_id}.parts"
    part_paths = sorted(parts_dir.glob("part-*.parquet"))
    if not part_paths:
        raise FileNotFoundError(f"No streaming parts found in {parts_dir}")
    if (staging_batch / "import_manifest.json").exists():
        raise FileExistsError(f"Housing import batch is already complete: {batch_id}")

    key_hashes: list[np.ndarray] = []
    key_ready_parts: list[np.ndarray] = []
    id_hashes: list[np.ndarray] = []
    part_rows: list[int] = []
    for part_path in part_paths:
        frame = pd.read_parquet(
            part_path, columns=["price_type", "source_record_id", *TRANSACTION_DEDUP_COLUMNS]
        )
        ready = (
            frame["price_type"].eq("transaction")
            & frame["city_key"].notna()
            & frame["community_name"].notna()
            & frame["deal_date"].notna()
            & frame["unit_price_cny_m2"].notna()
            & frame["building_area_m2"].notna()
            & frame["lon"].notna()
            & frame["lat"].notna()
        )
        key_hashes.append(
            pd.util.hash_pandas_object(
                frame[list(TRANSACTION_DEDUP_COLUMNS)], index=False
            ).to_numpy(dtype=np.uint64)
        )
        key_ready_parts.append(ready.to_numpy(dtype=bool))
        id_hashes.append(
            pd.util.hash_pandas_object(frame["source_record_id"], index=False).to_numpy(
                dtype=np.uint64
            )
        )
        part_rows.append(len(frame))
    all_hashes = np.concatenate(key_hashes)
    all_ready = np.concatenate(key_ready_parts)
    duplicate_key = pd.Series(all_hashes).duplicated(keep=False).to_numpy() & all_ready
    all_id_hashes = np.concatenate(id_hashes)
    duplicate_source_id = pd.Series(all_id_hashes).duplicated(keep=False).to_numpy()

    raw_batch.mkdir(parents=True, exist_ok=True)
    staging_batch.mkdir(parents=True, exist_ok=True)
    output_path = staging_batch / "housing_observations.parquet"
    writer: pq.ParquetWriter | None = None
    quality_counts: Counter[str] = Counter()
    offset = 0
    try:
        for part_path in part_paths:
            frame = pd.read_parquet(part_path)
            part_duplicate = duplicate_key[offset : offset + len(frame)]
            if part_duplicate.any():
                frame.loc[part_duplicate, "quality_flags"] = frame.loc[
                    part_duplicate, "quality_flags"
                ].map(lambda value: _append_flag_once(value, "duplicate_transaction_key"))
            part_dup_id = duplicate_source_id[offset : offset + len(frame)]
            if part_dup_id.any():
                frame.loc[part_dup_id, "quality_flags"] = frame.loc[
                    part_dup_id, "quality_flags"
                ].map(lambda value: _append_flag_once(value, "duplicate_source_record_id"))
            quality_counts.update(frame["quality_flags"].value_counts().to_dict())
            table = pa.Table.from_pandas(frame, preserve_index=False)
            if writer is None:
                writer = pq.ParquetWriter(output_path, table.schema)
            else:
                table = table.cast(writer.schema)
            writer.write_table(table)
            offset += len(frame)
    finally:
        if writer is not None:
            writer.close()

    source_hash = sha256_file(input_path)
    shutil.copy2(input_path, raw_batch / input_path.name)
    shutil.copy2(mapping_path, raw_batch / "import_mapping.yaml")
    manifest = {
        "schema": "housing_authorized_import_manifest_v1",
        "observation_schema": SCHEMA_VERSION,
        "pipeline_version": PIPELINE_VERSION,
        "import_mode": "streaming_xlsx_resumed",
        "chunk_rows": max(part_rows),
        "part_count": len(part_paths),
        "batch_id": batch_id,
        "metadata": mapping["metadata"],
        "input_file": input_path.name,
        "input_sha256": source_hash,
        "input_rows": int(sum(part_rows)),
        "output_rows": int(sum(part_rows)),
        "quality_flag_counts": dict(quality_counts),
        "output_file": str(output_path),
    }
    manifest_path = staging_batch / "import_manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    shutil.rmtree(parts_dir)
    return output_path, manifest_path


def import_authorized_export(
    input_path: Path,
    mapping_path: Path,
    raw_root: Path,
    staging_root: Path,
) -> tuple[Path, Path]:
    input_path = input_path.resolve()
    mapping = load_mapping(mapping_path)
    batch_id = str(mapping["metadata"]["batch_id"])
    raw_batch = raw_root / batch_id
    staging_batch = staging_root / batch_id
    if raw_batch.exists() or staging_batch.exists():
        raise FileExistsError(f"Housing import batch already exists: {batch_id}")

    source_hash = sha256_file(input_path)
    mapped_columns = (
        sorted(
            set(mapping.get("columns", {}).values())
            | _date_part_source_columns(mapping)
            | {str(rule["source_column"]) for rule in mapping.get("quality_flag_rules", [])}
        )
        or None
    )
    source = _load_table(
        input_path,
        columns=mapped_columns,
        sheet_name=mapping.get("sheet_name"),
    )
    normalized = normalize_authorized_export(source, mapping, source_hash)

    raw_batch.mkdir(parents=True)
    staging_batch.mkdir(parents=True)
    raw_copy = raw_batch / input_path.name
    shutil.copy2(input_path, raw_copy)
    mapping_copy = raw_batch / "import_mapping.yaml"
    shutil.copy2(mapping_path, mapping_copy)
    output_path = staging_batch / "housing_observations.parquet"
    normalized.to_parquet(output_path, index=False)

    manifest = {
        "schema": "housing_authorized_import_manifest_v1",
        "observation_schema": SCHEMA_VERSION,
        "pipeline_version": PIPELINE_VERSION,
        "batch_id": batch_id,
        "metadata": mapping["metadata"],
        "input_file": input_path.name,
        "input_sha256": source_hash,
        "input_rows": int(len(source)),
        "output_rows": int(len(normalized)),
        "quality_flag_counts": normalized["quality_flags"].value_counts().to_dict(),
        "output_file": str(output_path),
    }
    manifest_path = staging_batch / "import_manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    return output_path, manifest_path
